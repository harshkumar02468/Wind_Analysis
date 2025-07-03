import os
import pandas as pd
import numpy as np
from flask import Flask, render_template, request, jsonify, send_from_directory
from datetime import datetime, timedelta
from scipy.stats import weibull_min
from concurrent.futures import ThreadPoolExecutor, as_completed, ProcessPoolExecutor
import math
import atexit
import json
import traceback
import re
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from functools import lru_cache
import warnings
import multiprocessing
from multiprocessing import freeze_support, cpu_count, Manager
from numba import jit, njit, prange
import gc
import psutil
import logging
from logging.handlers import RotatingFileHandler
import time
from collections import defaultdict
import threading
from queue import Queue
import concurrent.futures
from werkzeug.serving import WSGIRequestHandler
import csv
import glob
app = Flask(__name__)
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
KML_DIR = os.path.join(BASE_DIR, 'static', 'kml')
@app.route("/")
def map():
    return render_template("map.html")
@app.route('/api/kml-files')
def list_kml_files():
    """Endpoint to list all available KMZ/KML files in the directory"""
    files = []
    logger.debug(f"Scanning directory for KMZ/KML files: {KML_DIR}")
    if not os.path.exists(KML_DIR):
        logger.error(f"Directory does not exist: {KML_DIR}")
        return jsonify({"error": "KMZ/KML directory not found"}), 404
    try:
        for filename in sorted(os.listdir(KML_DIR)):
            lower_filename = filename.lower()
            if lower_filename.endswith('.kmz') or lower_filename.endswith('.kml'):
                file_path = os.path.join(KML_DIR, filename)
                if os.path.isfile(file_path):
                    display_name = os.path.splitext(filename)[0].replace('_', ' ')
                    files.append({
                        'name': display_name,
                        'filename': filename,
                        'path': f'/static/kml/{filename}',
                        'type': 'kmz' if lower_filename.endswith('.kmz') else 'kml',
                        'size': os.path.getsize(file_path)
                    })
                    logger.debug(f"Found file: {filename} (Size: {os.path.getsize(file_path)} bytes)")
    except Exception as e:
        logger.error(f"Error scanning directory: {str(e)}")
        return jsonify({
            "error": "Error scanning directory",
            "details": str(e)
        }), 500
    logger.debug(f"Returning {len(files)} files")
    return jsonify(files)
@app.route('/static/kml/<path:filename>')
def serve_kml_file(filename):
    """Endpoint to serve KMZ/KML files"""
    try:
        logger.debug(f"Serving file: {filename}")
        return send_from_directory(KML_DIR, filename)
    except FileNotFoundError:
        logger.error(f"File not found: {filename}")
        return jsonify({"error": "File not found"}), 404
    except Exception as e:
        logger.error(f"Error serving file {filename}: {str(e)}")
        return jsonify({
            "error": "Error serving file",
            "details": str(e)
        }), 500
@app.route('/api/mast-summary', methods=['POST'])
def get_mast_summary():
    """Endpoint to get mast summary data from CSV files"""
    try:
        data = request.get_json()
        if not data:
            raise ValueError("No JSON data received")
        mast_code = data.get('mast_code', '').strip()
        if not mast_code:
            raise ValueError("Mast code not provided")
        clean_mast_code = ''.join(filter(str.isdigit, mast_code))
        logger.debug(f"Searching for mast code: {clean_mast_code}")
        csv_dir = os.path.join(BASE_DIR, 'static', 'CSV_files')
        csv_files = glob.glob(os.path.join(csv_dir, '*.csv'))
        if not csv_files:
            raise ValueError("No CSV files found in the CSV directory")
        mast_data = None
        for csv_file in csv_files:
            try:
                with open(csv_file, mode='r', encoding='utf-8-sig') as file:
                    reader = csv.DictReader(file)
                    for row in reader:
                        if ''.join(filter(str.isdigit, row.get('Mast Code', ''))) == clean_mast_code:
                            mast_data = {
                                "Mast Code": row.get('Mast Code', ''),
                                "Name of the Met Mast": row.get('Name of the Met Mast', ''),
                                "Mast Height(M)": row.get('Mast Height(M)', ''),
                                "Commissioning Date": row.get('Commissioning Date', ''),
                                "Date of Dismantled": row.get('Date of Dismantled', ''),
                                "Data Period": row.get('Data Period', ''),
                                "Zone": row.get('Zone', ''),
                                "Easting": row.get('Easting', ''),
                                "Northing": row.get('Northing', ''),
                                "Status": row.get('Status', ''),
                            }
                            break
                    if mast_data:
                        break
            except Exception as e:
                logger.error(f"Error reading CSV file {csv_file}: {str(e)}")
                continue
        if not mast_data:
            return jsonify({
                "error": f"Mast with code {clean_mast_code} not found in any CSV file"
            }), 404
        return jsonify(mast_data)
    except Exception as e:
        logger.error(f"Error in mast summary: {str(e)}")
        return jsonify({
            "error": str(e)
        }), 400
Speed_DATA_DIR = r"Z:\Data Analysis Team\Wind tool\data_harsh"
@app.route('/index_wind')
def index_wind():
    return render_template('index_wind.html')
@app.route('/get_mast', methods=['POST'])
def get_mast():
    data = request.get_json()
    state = data['state']
    try:
        state_dir = os.path.join(Speed_DATA_DIR, state)
        print(f"Looking for CSVs in: {state_dir}")  
        if not os.path.isdir(state_dir):
            return jsonify({
                'success': False,
                'message': f"Directory not found: {state_dir}"
            })
        csv_files = []
        for f in os.listdir(state_dir):
            if f.lower().endswith('.csv') and not f.startswith('.'):
                csv_files.append(os.path.join(state_dir, f))
        print(f"Found {len(csv_files)} CSV files")  
        if not csv_files:
            return jsonify({
                'success': False,
                'message': f"No CSV files found in {state_dir}"
            })
        masts = [os.path.splitext(os.path.basename(f))[0] for f in csv_files]
        print(f"Masts found: {masts}")  
        return jsonify({
            'success': True,
            'masts': sorted(masts)
        })
    except Exception as e:
        print(f"Error in get_mast: {str(e)}")  
        return jsonify({
            'success': False,
            'message': f"Server error: {str(e)}"
        })
@app.route('/get_column', methods=['POST'])
def get_column():
    data = request.get_json()
    state = data['state']
    mast = data['mast']
    csv_path = os.path.join(Speed_DATA_DIR, state, f"{mast}.csv")
    try:
        if not os.path.exists(csv_path):
            return jsonify({
                'success': False,
                'message': f"File not found: {mast}.csv"
            })
        df = pd.read_csv(csv_path, nrows=0)
        speed_columns = [col for col in df.columns 
                        if re.match(r'^Spd.*\[m/s\]$', col, re.IGNORECASE)]
        direction_columns = [col for col in df.columns 
                           if re.match(r'^Dir.*[°Â°]\]$', col, re.IGNORECASE)]
        if not speed_columns and not direction_columns:
            return jsonify({
                'success': False,
                'message': "No speed or direction columns found in the file"
            })
        return jsonify({
            'success': True,
            'speed_columns': sorted(speed_columns),
            'direction_columns': sorted(direction_columns)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        })
@app.route('/generate_analysis', methods=['POST'])
def generate_analysis():
    data = request.get_json()
    start_date = data['start_date']
    end_date = data['end_date']
    selections = data['selections']
    speed_analysis = data.get('speed_analysis', False)
    direction_analysis = data.get('direction_analysis', False)
    sector_count = data.get('sector_count', 12) if direction_analysis else None
    try:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        results = {
            'success': True,
            'speed_results': None,
            'direction_results': None
        }
        if speed_analysis:
            results['speed_results'] = process_speed_analysis(start_dt, end_dt, selections)
        if direction_analysis:
            results['direction_results'] = process_direction_analysis(start_dt, end_dt, selections, sector_count)
        return jsonify(results)
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        })
def process_speed_analysis(start_dt, end_dt, selections):
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                  'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    speed_results = {
        'monthly_data': [],
        'overall_averages': [],
        'selections': []
    }
    for month_num in range(1, 13):
        speed_results['monthly_data'].append({
            'month': month_names[month_num - 1],
            'values': []
        })
    for selection in selections:
        if not selection['speed_column']:
            continue
        state = selection['state']
        mast = selection['mast']
        speed_column = selection['speed_column']
        speed_results['selections'].append(selection)
        csv_path = os.path.join(Speed_DATA_DIR, state, f"{mast}.csv")
        try:
            try:
                df = pd.read_csv(csv_path)
                timestamp_col = None
                timestamp_format = None
                possible_timestamp_cols = [
                    'Timestamp (UTC+05:30)',
                    'Timestamp (UTC+00:00)',
                    'Timestamp (UTC+05)' ,
                    'Timestamp',
                    'Date/Time',
                    'DateTime'
                ]
                for col in possible_timestamp_cols:
                    if col in df.columns:
                        timestamp_col = col
                        break
                if not timestamp_col:
                    raise ValueError("No recognized timestamp column found")
                try:
                    df['Timestamp'] = pd.to_datetime(df[timestamp_col], format='%d%m%Y %H%M')
                except:
                    try:
                        df['Timestamp'] = pd.to_datetime(df[timestamp_col])
                    except:
                        raise ValueError("Could not parse timestamp column")
                df = df[(df['Timestamp'] >= start_dt) & (df['Timestamp'] <= end_dt)]
                if df.empty:
                    raise ValueError(f"No data found in the selected date range for {mast}")
                df[speed_column] = pd.to_numeric(df[speed_column], errors='coerce')
                df[speed_column] = df[speed_column].replace(9999, pd.NA)
                df['Month'] = df['Timestamp'].dt.month
                monthly_avg = df.groupby('Month')[speed_column].mean()
                overall_avg = df[speed_column].mean()
                speed_results['overall_averages'].append(overall_avg if not pd.isna(overall_avg) else 0)
                for month_num in range(1, 13):
                    month_avg = monthly_avg.get(month_num, 0)  
                    speed_results['monthly_data'][month_num - 1]['values'].append(
                        month_avg if not pd.isna(month_avg) else 0
                    )
            except Exception as e:
                print(f"Error processing {state}/{mast}: {str(e)}")
                raise
        except Exception as e:
            print(f"Error processing {state}/{mast}: {str(e)}")
            speed_results['overall_averages'].append(0)
            for month_data in speed_results['monthly_data']:
                month_data['values'].append(0)
            continue
    return speed_results
def process_direction_analysis(start_dt, end_dt, selections, sector_count):
    sector_angle = 360 / sector_count
    sector_midpoints = np.arange(0, 360, sector_angle)
    sector_labels = []
    for i in range(sector_count):
        start_angle = i * sector_angle
        end_angle = (i + 1) * sector_angle
        sector_labels.append(f"{int(start_angle)}°-{int(end_angle)}°")
    direction_results = {
        'frequency_data': [],
        'sector_midpoints': sector_midpoints.tolist(),
        'sector_labels': sector_labels,
        'selections': []
    }
    for selection in selections:
        if not selection['direction_column']:
            continue
        state = selection['state']
        mast = selection['mast']
        direction_column = selection['direction_column']
        direction_results['selections'].append(selection)
        csv_path = os.path.join(Speed_DATA_DIR, state, f"{mast}.csv")
        try:
            try:
                df = pd.read_csv(csv_path)
                timestamp_col = None
                possible_timestamp_cols = [
                    'Timestamp (UTC+05:30)',
                    'Timestamp (UTC+00:00)',
                    'Timestamp (UTC+05)',
                    'Timestamp',
                    'Date/Time',
                    'DateTime'
                ]
                for col in possible_timestamp_cols:
                    if col in df.columns:
                        timestamp_col = col
                        break
                if not timestamp_col:
                    raise ValueError("No recognized timestamp column found")
                try:
                    df['Timestamp'] = pd.to_datetime(df[timestamp_col], format='%d%m%Y %H%M')
                except:
                    try:
                        df['Timestamp'] = pd.to_datetime(df[timestamp_col])
                    except:
                        raise ValueError("Could not parse timestamp column")
                df = df[(df['Timestamp'] >= start_dt) & (df['Timestamp'] <= end_dt)]
                if df.empty:
                    raise ValueError(f"No data found in the selected date range for {mast}")
                df[direction_column] = pd.to_numeric(df[direction_column], errors='coerce')
                df[direction_column] = df[direction_column].replace(9999, pd.NA)
                df = df[~pd.isna(df[direction_column])]
                if df.empty:
                    raise ValueError(f"No valid direction data found for {mast}")
                df['Sector'] = (df[direction_column] // sector_angle) % sector_count
                sector_counts = df['Sector'].value_counts().sort_index()
                total_count = len(df)
                sector_frequencies = []
                for sector in range(sector_count):
                    count = sector_counts.get(sector, 0)
                    frequency = (count / total_count) * 100
                    midpoint = sector * sector_angle + (sector_angle / 2)
                    sector_frequencies.append({
                        'sector': sector,
                        'midpoint': midpoint,
                        'frequency': frequency,
                        'label': sector_labels[sector]
                    })
                direction_results['frequency_data'].append(sector_frequencies)
            except Exception as e:
                print(f"Error processing {state}/{mast}: {str(e)}")
                raise
        except Exception as e:
            print(f"Error processing {state}/{mast}: {str(e)}")
            sector_frequencies = []
            for sector in range(sector_count):
                midpoint = sector * sector_angle + (sector_angle / 2)
                sector_frequencies.append({
                    'sector': sector,
                    'midpoint': midpoint,
                    'frequency': 0,
                    'label': sector_labels[sector]
                })
            direction_results['frequency_data'].append(sector_frequencies)
            continue
    return direction_results
BASE_DIR = r"Z:\Data Analysis Team\Wind tool"
POWER_CURVES_DIR = os.path.join(BASE_DIR, "power_curves")
DATA_DIR = os.path.join(BASE_DIR, "data_harsh")
EXTRAPOLATED_DATA_PATH = os.path.join(BASE_DIR, "extrapolated_data.json")
ERROR_LOG_PATH = os.path.join(BASE_DIR, "error_logs.json")
CACHE_DIR = os.path.join(BASE_DIR, "cache")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        RotatingFileHandler('wind_analysis.log', maxBytes=10*1024*1024, backupCount=5),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)
os.makedirs(CACHE_DIR, exist_ok=True)
loss_factors = {
    "machineAvailability": 5.0,
    "gridLosses": 2.0,
    "transmissionLosses": 3.0,
    "Airdensitycorrection": 9.5,
    "Horizontalextrapolation": 2.0,
    "otherLosses": 1.0
}
extrapolated_data = {}
error_tracker = {
    'last_errors': [],
    'error_counts': {},
    'common_failures': []
}
processed_data_cache = {}
data_sheet_cache = {}
MONTHS = ["January", "February", "March", "April", "May", "June", 
          "July", "August", "September", "October", "November", "December", "Average"]
excel_preparation_queue = Queue()
excel_preparation_cache = {}
excel_preparation_lock = threading.Lock()
worker_pool = ThreadPoolExecutor(max_workers=cpu_count() * 2)
def log_error(error_type, details):
    """Log errors with detailed context"""
    error_entry = {
        'timestamp': datetime.now().isoformat(),
        'type': error_type,
        'details': details,
        'traceback': traceback.format_exc()
    }
    error_tracker['last_errors'].append(error_entry)
    if len(error_tracker['last_errors']) > 100:
        error_tracker['last_errors'].pop(0)
    error_tracker['error_counts'][error_type] = error_tracker['error_counts'].get(error_type, 0) + 1
    logger.error(f"{error_type}: {details}\n{traceback.format_exc()}")
    try:
        with open(ERROR_LOG_PATH, 'w') as f:
            json.dump(error_tracker, f, indent=2)
    except Exception as e:
        logger.error(f"Failed to save error log: {str(e)}")
def get_safe_chunksize(file_size):
    """Calculate safe chunksize based on available memory"""
    try:
        available_mem = psutil.virtual_memory().available * 0.8  
        estimated_mem_per_chunk = file_size / 100000  
        safe_chunksize = max(10000, min(100000, int(available_mem / estimated_mem_per_chunk)))
        logger.debug(f"Calculated chunksize: {safe_chunksize} (File: {file_size/1024/1024:.2f}MB, Available mem: {available_mem/1024/1024:.2f}MB)")
        return safe_chunksize
    except Exception as e:
        log_error("CHUNKSIZE_CALCULATION_ERROR", f"Failed to calculate chunksize: {str(e)}")
        return 100000  
def update_total_loss_factor():
    global total_loss_factor
    try:
        total_loss_factor = (
            (100 - loss_factors["machineAvailability"]) / 100 *
            (100 - loss_factors["gridLosses"]) / 100 *
            (100 - loss_factors["transmissionLosses"]) / 100 *
            (100 - loss_factors["Airdensitycorrection"]) / 100 *
            (100 - loss_factors["Horizontalextrapolation"]) / 100 *
            (100 - loss_factors["otherLosses"]) / 100
        )
        logger.debug(f"Updated total loss factor: {total_loss_factor}")
    except Exception as e:
        log_error("LOSS_FACTOR_ERROR", f"Failed to calculate total loss factor: {str(e)}")
        total_loss_factor = 0.8  
update_total_loss_factor()
def initialize_extrapolated_file():
    global extrapolated_data
    try:
        if os.path.exists(EXTRAPOLATED_DATA_PATH):
            with open(EXTRAPOLATED_DATA_PATH, 'w') as f:
                f.write('{}')
        extrapolated_data = {}
        logger.info("Initialized empty extrapolated data store")
    except Exception as e:
        logger.error(f"Error initializing extrapolated data: {str(e)}")
        traceback.print_exc()
        extrapolated_data = {}
def clear_extrapolated_data():
    try:
        if os.path.exists(EXTRAPOLATED_DATA_PATH):
            with open(EXTRAPOLATED_DATA_PATH, 'w') as f:
                f.write('{}')
            logger.info("Extrapolated data file cleared successfully")
        else:
            logger.info("No extrapolated data file to clear")
    except Exception as e:
        logger.error(f"Error clearing extrapolated data: {str(e)}")
        traceback.print_exc()
def save_extrapolated_data():
    try:
        global extrapolated_data
        extrapolated_data = {k: v for k, v in extrapolated_data.items() if v}
        with open(EXTRAPOLATED_DATA_PATH, 'w') as f:
            json.dump(extrapolated_data, f, indent=2)
        logger.info("Extrapolated data saved successfully")
        return True
    except Exception as e:
        logger.error(f"Error saving extrapolated data: {str(e)}")
        traceback.print_exc()
        return False
initialize_extrapolated_file()
atexit.register(clear_extrapolated_data)
class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (np.integer, np.int32, np.int64)):
            return int(obj)
        elif isinstance(obj, (np.floating, np.float32, np.float64)):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, np.bool_):
            return bool(obj)
        return super(NumpyEncoder, self).default(obj)
app.json_encoder = NumpyEncoder
@njit
def power_interpolation(wind_speeds, power_outputs, speed):
    """Optimized power interpolation using binary search"""
    if speed <= wind_speeds[0]:
        return 0.0
    if speed >= wind_speeds[-1]:
        return power_outputs[-1]
    left = 0
    right = len(wind_speeds) - 1
    while left <= right:
        mid = (left + right) // 2
        if wind_speeds[mid] < speed:
            left = mid + 1
        elif wind_speeds[mid] > speed:
            right = mid - 1
        else:
            return power_outputs[mid]
    x0, x1 = wind_speeds[right], wind_speeds[left]
    y0, y1 = power_outputs[right], power_outputs[left]
    return y0 + (y1 - y0) * (speed - x0) / (x1 - x0)
@njit(parallel=True)
def batch_power_interpolation(wind_speeds, power_outputs, speeds):
    """Process multiple wind speeds in parallel"""
    results = np.empty(len(speeds))
    for i in prange(len(speeds)):
        results[i] = power_interpolation(wind_speeds, power_outputs, speeds[i])
    return results
@lru_cache(maxsize=32)
def get_power_curves_cached():
    return [f.replace('.csv', '') for f in os.listdir(POWER_CURVES_DIR) if f.endswith('.csv')]
@lru_cache(maxsize=32)
def get_state_masts_cached(state):
    state_dir = os.path.join(DATA_DIR, state)
    if not os.path.exists(state_dir):
        return []
    return [f.replace('.csv', '') for f in os.listdir(state_dir) if f.endswith('.csv')]
@lru_cache(maxsize=128)
def get_mast_columns_cached(state, mast):
    file_path = os.path.join(DATA_DIR, state, f"{mast}.csv")
    if not os.path.exists(file_path):
        return []
    try:
        with open(file_path, 'r') as f:
            first_line = f.readline()
        return first_line.strip().split(',')
    except Exception as e:
        logger.error(f"Error reading columns for {state}/{mast}: {str(e)}")
        return []
def get_power_curves():
    return get_power_curves_cached()
def get_state_masts(state):
    return get_state_masts_cached(state)
def get_mast_columns(state, mast):
    return get_mast_columns_cached(state, mast)
@njit
def clean_numeric_value_numba(value):
    if np.isnan(value) or value == 9999:
        return np.nan
    return value
def clean_numeric_value(value):
    if pd.isna(value) or value == 9999:
        return np.nan
    if isinstance(value, str):
        try:
            return float(value.replace(',', ''))
        except ValueError:
            return np.nan
    return float(value)
@njit
def calculate_wind_shear_numba(v1, v2, h1, h2):
    if v1 <= 0 or v2 <= 0 or h1 <= 0 or h2 <= 0:
        return np.nan
    try:
        return math.log(v2 / v1) / math.log(h2 / h1)
    except:
        return np.nan
def calculate_wind_shear(v1, v2, h1, h2):
    try:
        if v1 <= 0 or v2 <= 0 or h1 <= 0 or h2 <= 0:
            return None
        return math.log(v2 / v1) / math.log(h2 / h1)
    except Exception as e:
        logger.error(f"Error calculating wind shear: {str(e)}")
        traceback.print_exc()
        return None
def extrapolate_wind_speeds(df, source_col, target_height, alpha=None):
    try:
        ref_height = float(source_col.split(' ')[1].replace('m', ''))
        if alpha is None:
            alpha = 1/7
        new_col = f"Spd {target_height}m (extrapolated from {ref_height}m)"
        df[new_col] = df[source_col] * (target_height / ref_height) ** alpha
        return new_col
    except Exception as e:
        logger.error(f"Error in extrapolation: {str(e)}")
        traceback.print_exc()
        return None
def load_extrapolated_data(state, mast, speed_col):
    key = f"{state}_{mast}_{speed_col}"
    return extrapolated_data.get(key, {})
def apply_loss_factors(gross_energy):
    return gross_energy * total_loss_factor
@lru_cache(maxsize=32)
def load_power_curve(power_curve_name):
    power_curve_path = os.path.join(POWER_CURVES_DIR, f"{power_curve_name}.csv")
    if os.path.exists(power_curve_path):
        try:
            pc_df = pd.read_csv(power_curve_path, thousands=',', 
                              dtype={'Wind Speed (m/s)': np.float32, 
                                    'Power Output (kW)': np.float32})
            if 'Wind Speed (m/s)' in pc_df.columns and 'Power Output (kW)' in pc_df.columns:
                pc_df['Power Output (kW)'] = pc_df['Power Output (kW)'].apply(clean_numeric_value)
                return {
                    'wind_speeds': pc_df['Wind Speed (m/s)'].values,
                    'power_outputs': pc_df['Power Output (kW)'].values,
                    'rated_power': pc_df['Power Output (kW)'].iloc[-1]
                }
        except Exception as e:
            logger.error(f"Error loading power curve {power_curve_name}: {str(e)}")
    return None
def calculate_power(wind_speed, power_curve):
    if pd.isna(wind_speed) or wind_speed <= 0:
        return 0
    return power_interpolation(power_curve['wind_speeds'], 
                             power_curve['power_outputs'], 
                             wind_speed)
@njit
def calculate_energy_pattern_factor_numba(wind_speeds):
    if len(wind_speeds) == 0:
        return np.nan
    mean_speed = np.mean(wind_speeds)
    if mean_speed == 0:
        return np.nan
    cube_sum = np.sum(wind_speeds**3)
    return cube_sum / (len(wind_speeds) * mean_speed**3)
def calculate_energy_pattern_factor(wind_speeds):
    wind_speeds = np.array([x for x in wind_speeds if not pd.isna(x)])
    if len(wind_speeds) == 0:
        return np.nan
    return calculate_energy_pattern_factor_numba(wind_speeds)
def clean_column_name(col_name):
    if col_name.startswith('Spd '):
        parts = col_name.split(' ')
        height = parts[1].replace('m', '')
        direction = parts[2] if len(parts) > 2 else ''
        direction = direction.replace('[', '').replace(']', '').replace('m/s', '')
        if direction in ['A', 'B']:
            direction = 'N' if direction == 'A' else 'S'
        elif not direction or direction == '[m/s]':
            direction = ''
        return f"{height} {direction}".strip()
    elif col_name.startswith('Dir '):
        parts = col_name.split(' ')
        height = parts[1].replace('m', '')
        direction = parts[2] if len(parts) > 2 else ''
        direction = direction.replace('[', '').replace(']', '').replace('°', '')
        return f"{height} {direction}".strip()
    elif col_name.startswith('Tmp ') or col_name.startswith('Air Density '):
        height = col_name.split(' ')[1].replace('m', '')
        return height
    return col_name
def get_height_from_column(col_name):
    if col_name.startswith('Spd '):
        try:
            height = int(col_name.split(' ')[1].replace('m', ''))
            direction = col_name.split(' ')[2] if len(col_name.split(' ')) > 2 else ''
            return height, direction
        except:
            return 0, ''
    return 0, ''
def is_relevant_height(col_name):
    """This function is no longer used to restrict EPF calculations"""
    height, direction = get_height_from_column(col_name)
    return True  
def validate_file(state, mast):
    """Validate the structure of a data file"""
    try:
        file_path = os.path.join(DATA_DIR, state, f"{mast}.csv")
        if not os.path.exists(file_path):
            return {'valid': False, 'error': 'File not found', 'file_path': file_path}
        if os.path.getsize(file_path) == 0:
            return {'valid': False, 'error': 'Empty file', 'file_path': file_path}
        with open(file_path, 'r') as f:
            header = f.readline()
            sample_data = f.readline()
        if not header:
            return {'valid': False, 'error': 'Empty header', 'file_path': file_path}
        if not sample_data:
            return {'valid': False, 'error': 'No data rows', 'file_path': file_path}
        headers = header.strip().split(',')
        timestamp_col = next((col for col in headers if 'Timestamp' in col), None)
        if not timestamp_col:
            return {'valid': False, 'error': 'No timestamp column found', 'headers': headers, 'file_path': file_path}
        return {
            'valid': True,
            'file_path': file_path,
            'headers': headers,
            'timestamp_col': timestamp_col,
            'sample_data': sample_data.strip()
        }
    except Exception as e:
        return {'valid': False, 'error': str(e), 'traceback': traceback.format_exc()}
def process_data_chunk(chunk_df, speed_col, power_curve, date_range, state, mast):
    """Process a chunk of data with enhanced error handling and optimized operations"""
    try:
        if speed_col not in chunk_df.columns:
            log_error("MISSING_SPEED_COLUMN", f"Speed column {speed_col} not found in data")
            return None
        if 'Timestamp' not in chunk_df.columns:
            log_error("MISSING_TIMESTAMP", "Timestamp column not found in chunk")
            return None
        chunk_df[speed_col] = pd.to_numeric(chunk_df[speed_col], errors='coerce')
        chunk_df[speed_col] = chunk_df[speed_col].replace(9999, np.nan)
        valid_rows = chunk_df[~pd.isna(chunk_df[speed_col])]
        if len(valid_rows) == 0:
            logger.debug(f"No valid wind speed data in chunk for {state}/{mast}")
            return None
        if date_range['type'] == 'single_day':
            target_date = datetime.strptime(date_range['value'], '%d-%m-%Y').date()
            filtered_df = valid_rows[valid_rows['Timestamp'].dt.date == target_date]
        elif date_range['type'] == 'monthly':
            month, year = map(int, date_range['value'].split('-'))
            filtered_df = valid_rows[
                (valid_rows['Timestamp'].dt.month == month) & 
                (valid_rows['Timestamp'].dt.year == year)
            ]
        elif date_range['type'] == 'yearly':
            year = int(date_range['value'])
            filtered_df = valid_rows[valid_rows['Timestamp'].dt.year == year]
        elif date_range['type'] == 'date_range':
            start_date = datetime.strptime(date_range['start'], '%d-%m-%Y').date()
            end_date = datetime.strptime(date_range['end'], '%d-%m-%Y').date()
            filtered_df = valid_rows[
                (valid_rows['Timestamp'].dt.date >= start_date) & 
                (valid_rows['Timestamp'].dt.date <= end_date)
            ]
        elif date_range['type'] == 'average':
            filtered_df = valid_rows.copy()
        if len(filtered_df) == 0:
            logger.debug(f"No data after date filtering for {state}/{mast}")
            return None
        wind_speeds = filtered_df[speed_col].values
        avg_wind_speed = np.mean(wind_speeds)
        std_dev = np.std(wind_speeds)
        try:
            if len(wind_speeds) > 10:
                shape, loc, scale = weibull_min.fit(wind_speeds, floc=0)
            else:
                shape, scale = 2, 8  
        except Exception as e:
            log_error("WEIBULL_FIT_ERROR", f"Weibull fit failed for {state}/{mast}: {str(e)}")
            shape, scale = 2, 8
        power_output = batch_power_interpolation(
            power_curve['wind_speeds'],
            power_curve['power_outputs'],
            wind_speeds
        )
        avg_power = np.mean(power_output) if len(power_output) > 0 else 0
        if date_range['type'] == 'monthly':
            month, year = map(int, date_range['value'].split('-'))
            days_in_month = (datetime(year, month + 1, 1) - datetime(year, month, 1)).days if month < 12 else 31
            total_hours = days_in_month * 24
        elif date_range['type'] == 'yearly':
            total_hours = 365 * 24
        elif date_range['type'] == 'date_range':
            days = (end_date - start_date).days + 1
            total_hours = days * 24
        else:  
            total_hours = 24
        gross_energy = avg_power * total_hours if avg_power else 0
        net_energy = apply_loss_factors(gross_energy) if gross_energy else 0
        rated_power = power_curve['rated_power']
        gross_plf = (gross_energy / (rated_power * total_hours)) * 100 if rated_power > 0 and total_hours > 0 else 0
        net_plf = (net_energy / (rated_power * total_hours)) * 100 if rated_power > 0 and total_hours > 0 else 0
        monthly_stats = []
        if date_range['type'] in ['yearly', 'date_range', 'average']:
            grouped = filtered_df.groupby([filtered_df['Timestamp'].dt.year, filtered_df['Timestamp'].dt.month])
            for (year, month), month_df in grouped:
                month_wind_speeds = month_df[speed_col].values
                month_avg_speed = np.mean(month_wind_speeds)
                month_power = batch_power_interpolation(
                    power_curve['wind_speeds'],
                    power_curve['power_outputs'],
                    month_wind_speeds
                )
                month_avg_power = np.mean(month_power)
                if month == 2 and year % 4 == 0 and (year % 100 != 0 or year % 400 == 0):
                    days_in_month = 29
                elif month == 2:
                    days_in_month = 28
                elif month in [4, 6, 9, 11]:
                    days_in_month = 30
                else:
                    days_in_month = 31
                total_hours = days_in_month * 24
                gross_energy = month_avg_power * total_hours
                net_energy = apply_loss_factors(gross_energy)
                gross_plf = (gross_energy / (rated_power * total_hours)) * 100 if rated_power > 0 and total_hours > 0 else 0
                net_plf = (net_energy / (rated_power * total_hours)) * 100 if rated_power > 0 and total_hours > 0 else 0
                monthly_stats.append({
                    'year': year,
                    'month': month,
                    'month_name': datetime(year, month, 1).strftime('%B'),
                    'avg_wind_speed': month_avg_speed,
                    'avg_power': month_avg_power,
                    'gross_energy': gross_energy,
                    'net_energy': net_energy,
                    'gross_plf': gross_plf,
                    'net_plf': net_plf
                })
        return {
            'avg_wind_speed': round(avg_wind_speed, 2) if not pd.isna(avg_wind_speed) else 0,
            'std_dev': round(std_dev, 2) if not pd.isna(std_dev) else 0,
            'weibull_shape': round(shape, 2) if not pd.isna(shape) else 0,
            'weibull_scale': round(scale, 2) if not pd.isna(scale) else 0,
            'mean_monthly_mean': round(avg_wind_speed, 2) if not pd.isna(avg_wind_speed) else 0,
            'avg_power': round(avg_power, 2) if not pd.isna(avg_power) else 0,
            'gross_energy': round(gross_energy, 2) if not pd.isna(gross_energy) else 0,
            'net_energy': round(net_energy, 2) if not pd.isna(net_energy) else 0,
            'gross_plf': round(gross_plf, 2) if not pd.isna(gross_plf) else 0,
            'net_plf': round(net_plf, 2) if not pd.isna(net_plf) else 0,
            'hours': total_hours,
            'valid_records': len(wind_speeds),
            'monthly_stats': monthly_stats
        }
    except Exception as e:
        log_error("CHUNK_PROCESSING_ERROR", f"Failed to process chunk for {state}/{mast}: {str(e)}")
        return None
def process_data(state, mast, speed_col, power_curve_name, date_range):
    """Main processing function with comprehensive error handling and optimization"""
    try:
        logger.info(f"Starting processing for {state}/{mast}/{speed_col} with {power_curve_name}")
        cache_key = f"{state}_{mast}_{speed_col}_{power_curve_name}_{date_range['type']}_{date_range.get('value', '')}_{date_range.get('start', '')}_{date_range.get('end', '')}"
        if cache_key in processed_data_cache:
            logger.info(f"Returning cached result for {cache_key}")
            return processed_data_cache[cache_key]
        if not all([state, mast, speed_col, power_curve_name]):
            raise ValueError("Missing required parameters")
        power_curve = load_power_curve(power_curve_name)
        if not power_curve:
            raise ValueError(f"Invalid power curve: {power_curve_name}")
        mast_path = os.path.join(DATA_DIR, state, f"{mast}.csv")
        if not os.path.exists(mast_path):
            raise ValueError(f"File not found: {mast_path}")
        file_size = os.path.getsize(mast_path)
        chunksize = get_safe_chunksize(file_size)
        logger.info(f"Processing {mast_path} (size: {file_size/1024/1024:.2f}MB) with chunksize {chunksize}")
        complete_years = set()
        if date_range['type'] == 'average':
            logger.info("Identifying complete years for average calculation")
            for chunk in pd.read_csv(mast_path, chunksize=chunksize):
                try:
                    cols = chunk.columns.tolist()
                    timestamp_col = next((c for c in cols if 'Timestamp' in c), None)
                    if not timestamp_col:
                        logger.warning("No timestamp column found in chunk")
                        continue
                    try:
                        chunk['Timestamp'] = pd.to_datetime(chunk[timestamp_col], format='%d%m%Y %H%M', errors='coerce')
                    except:
                        chunk['Timestamp'] = pd.to_datetime(chunk[timestamp_col], errors='coerce')
                    chunk = chunk.dropna(subset=['Timestamp'])
                    year_month_counts = chunk.groupby([
                        chunk['Timestamp'].dt.year,
                        chunk['Timestamp'].dt.month
                    ]).size()
                    yearly_counts = year_month_counts.groupby(level=0).count()
                    complete_years.update(yearly_counts[yearly_counts >= 12].index)
                except Exception as e:
                    log_error("YEAR_IDENTIFICATION_ERROR", f"Failed to process chunk for year identification: {str(e)}")
                finally:
                    del chunk
                    gc.collect()
            logger.info(f"Complete years identified: {complete_years}")
        results = []
        for chunk_idx, chunk in enumerate(pd.read_csv(mast_path, chunksize=chunksize)):
            try:
                logger.debug(f"Processing chunk {chunk_idx + 1}")
                cols = chunk.columns.tolist()
                timestamp_col = next((c for c in cols if 'Timestamp' in c), None)
                if not timestamp_col:
                    logger.warning("No timestamp column found in chunk")
                    continue
                try:
                    chunk['Timestamp'] = pd.to_datetime(chunk[timestamp_col], format='%d%m%Y %H%M', errors='coerce')
                except:
                    chunk['Timestamp'] = pd.to_datetime(chunk[timestamp_col], errors='coerce')
                chunk = chunk.dropna(subset=['Timestamp'])
                if date_range['type'] == 'average' and complete_years:
                    chunk = chunk[chunk['Timestamp'].dt.year.isin(complete_years)]
                chunk_result = process_data_chunk(chunk, speed_col, power_curve, date_range, state, mast)
                if chunk_result:
                    results.append(chunk_result)
            except Exception as e:
                log_error("CHUNK_PROCESSING_ERROR", f"Failed to process chunk {chunk_idx}: {str(e)}")
            finally:
                del chunk
                gc.collect()
        if not results:
            raise ValueError("No valid data after processing")
        valid_results = [r for r in results if r and r['valid_records'] > 0]
        if not valid_results:
            raise ValueError("No valid results after filtering")
        avg_wind_speed = np.nanmean([r['avg_wind_speed'] for r in valid_results])
        std_dev = np.nanmean([r['std_dev'] for r in valid_results])
        shape = np.nanmean([r['weibull_shape'] for r in valid_results])
        scale = np.nanmean([r['weibull_scale'] for r in valid_results])
        avg_power = np.nanmean([r['avg_power'] for r in valid_results])
        if date_range['type'] in ['single_day', 'monthly', 'yearly', 'date_range']:
            total_hours = valid_results[0]['hours']
            gross_energy = np.nansum([r['gross_energy'] for r in valid_results])
            net_energy = np.nansum([r['net_energy'] for r in valid_results])
        else:  
            total_hours = 0
            gross_energy = 0
            net_energy = 0
        gross_plf = np.nanmean([r['gross_plf'] for r in valid_results])
        net_plf = np.nanmean([r['net_plf'] for r in valid_results])
        combined_monthly_stats = []
        if date_range['type'] in ['yearly', 'date_range', 'average']:
            monthly_stats_dict = {}
            for result in valid_results:
                for month_data in result['monthly_stats']:
                    key = (month_data['year'], month_data['month'])
                    if key not in monthly_stats_dict:
                        monthly_stats_dict[key] = {
                            'year': month_data['year'],
                            'month': month_data['month'],
                            'month_name': month_data['month_name'],
                            'avg_wind_speed': [],
                            'avg_power': [],
                            'gross_energy': [],
                            'net_energy': [],
                            'gross_plf': [],
                            'net_plf': []
                        }
                    monthly_stats_dict[key]['avg_wind_speed'].append(month_data['avg_wind_speed'])
                    monthly_stats_dict[key]['avg_power'].append(month_data['avg_power'])
                    monthly_stats_dict[key]['gross_energy'].append(month_data['gross_energy'])
                    monthly_stats_dict[key]['net_energy'].append(month_data['net_energy'])
                    monthly_stats_dict[key]['gross_plf'].append(month_data['gross_plf'])
                    monthly_stats_dict[key]['net_plf'].append(month_data['net_plf'])
            for key, stats in monthly_stats_dict.items():
                combined_monthly_stats.append({
                    'year': stats['year'],
                    'month': stats['month'],
                    'month_name': stats['month_name'],
                    'avg_wind_speed': np.nanmean(stats['avg_wind_speed']),
                    'avg_power': np.nanmean(stats['avg_power']),
                    'gross_energy': np.nansum(stats['gross_energy']),
                    'net_energy': np.nansum(stats['net_energy']),
                    'gross_plf': np.nanmean(stats['gross_plf']),
                    'net_plf': np.nanmean(stats['net_plf'])
                })
        clean_mast = re.sub(r'\(\d{2}-\d{4}\)', '', mast).strip()
        result = {
            'state': state,
            'mast': mast,
            'speed_col': speed_col,
            'power_curve': power_curve_name,
            'rated_power': power_curve['rated_power'],
            'avg_wind_speed': avg_wind_speed,
            'std_dev': std_dev,
            'weibull_shape': shape,
            'weibull_scale': scale,
            'mean_monthly_mean': avg_wind_speed,
            'gross_energy': gross_energy,
            'net_energy': net_energy,
            'gross_plf': gross_plf,
            'net_plf': net_plf,
            'hours': total_hours,
            'monthly_stats': combined_monthly_stats,
            'clean_mast': clean_mast,
            'mast_id': mast.split('-')[1] if '-' in mast else mast,
            'avg_power': avg_power
        }
        processed_data_cache[cache_key] = result
        logger.info(f"Successfully processed {state}/{mast} and cached result")
        return result
    except Exception as e:
        log_error("PROCESS_DATA_ERROR", f"Failed to process {state}/{mast}/{speed_col}: {str(e)}")
        return None
def calculate_data_sheet_stats(state, mast, date_range, power_curves):
    """Calculate detailed statistics for data sheet with proper error handling and optimization"""
    try:
        logger.info(f"Calculating data sheet stats for {state}/{mast}")
        cache_key = f"datasheet_{state}_{mast}_{date_range['type']}_{date_range.get('value', '')}_{date_range.get('start', '')}_{date_range.get('end', '')}_{'_'.join(power_curves)}"
        if cache_key in data_sheet_cache:
            logger.info(f"Returning cached data sheet for {cache_key}")
            return data_sheet_cache[cache_key]
        file_path = os.path.join(DATA_DIR, state, f"{mast}.csv")
        if not os.path.exists(file_path):
            raise ValueError(f"File not found: {file_path}")
        power_curve_dfs = {}
        for pc in power_curves:
            pc_data = load_power_curve(pc)
            if pc_data:
                power_curve_dfs[pc] = pc_data
        if not power_curve_dfs:
            raise ValueError("No valid power curves loaded")
        chunksize = get_safe_chunksize(os.path.getsize(file_path))
        all_results = []
        for chunk in pd.read_csv(file_path, chunksize=chunksize):
            try:
                cols = chunk.columns.tolist()
                timestamp_col = next((c for c in cols if 'Timestamp' in c), None)
                if not timestamp_col:
                    logger.warning("No timestamp column found in chunk")
                    continue
                try:
                    chunk['Timestamp'] = pd.to_datetime(chunk[timestamp_col], format='%d%m%Y %H%M', errors='coerce')
                except:
                    chunk['Timestamp'] = pd.to_datetime(chunk[timestamp_col], errors='coerce')
                chunk = chunk.dropna(subset=['Timestamp'])
                if date_range['type'] == 'single_day':
                    target_date = datetime.strptime(date_range['value'], '%d-%m-%Y').date()
                    filtered_df = chunk[chunk['Timestamp'].dt.date == target_date]
                elif date_range['type'] == 'monthly':
                    month, year = map(int, date_range['value'].split('-'))
                    filtered_df = chunk[
                        (chunk['Timestamp'].dt.month == month) & 
                        (chunk['Timestamp'].dt.year == year)
                    ]
                elif date_range['type'] == 'yearly':
                    year = int(date_range['value'])
                    filtered_df = chunk[chunk['Timestamp'].dt.year == year]
                elif date_range['type'] == 'date_range':
                    start_date = datetime.strptime(date_range['start'], '%d-%m-%Y').date()
                    end_date = datetime.strptime(date_range['end'], '%d-%m-%Y').date()
                    filtered_df = chunk[
                        (chunk['Timestamp'].dt.date >= start_date) & 
                        (chunk['Timestamp'].dt.date <= end_date)
                    ]
                elif date_range['type'] == 'average':
                    filtered_df = chunk.copy()
                if filtered_df.empty:
                    continue
                speed_cols = [col for col in chunk.columns if col.startswith('Spd ')]
                dir_cols = [col for col in chunk.columns if col.startswith('Dir ')]
                temp_cols = [col for col in chunk.columns if col.startswith('Tmp ')]
                air_density_cols = [col for col in chunk.columns if col.startswith('Air Density ')]
                grouped = filtered_df.groupby([filtered_df['Timestamp'].dt.year, filtered_df['Timestamp'].dt.month])
                for group_key, month_df in grouped:
                    year, month = group_key
                    month_data = {
                        'year': int(year),
                        'month': int(month),
                        'month_name': datetime(year, month, 1).strftime('%B'),
                        'parameters': {},
                        'energy': []
                    }
                    for col in speed_cols:
                        clean_name = clean_column_name(col)
                        vals = month_df[col].replace(9999, np.nan).dropna()
                        if len(vals) > 0:
                            mean_val = np.mean(vals)
                            month_data['parameters'][f"Mean Wind Speed {clean_name} (m/s)"] = float(mean_val)
                    for col in dir_cols:
                        clean_name = clean_column_name(col)
                        directions = month_df[col].replace(9999, np.nan).dropna()
                        if len(directions) > 0:
                            sin_sum = np.sum(np.sin(np.radians(directions)))
                            cos_sum = np.sum(np.cos(np.radians(directions)))
                            mean_dir = np.degrees(np.arctan2(sin_sum, cos_sum)) % 360
                            month_data['parameters'][f"Mean Wind Direction {clean_name} (°)"] = float(mean_dir)
                    for col in temp_cols:
                        clean_name = clean_column_name(col)
                        max_temp = month_df[col].replace(9999, np.nan).max()
                        if not pd.isna(max_temp):
                            month_data['parameters'][f"Max Temperature {clean_name} (°C)"] = float(max_temp)
                    for col in air_density_cols:
                        clean_name = clean_column_name(col)
                        mean_density = month_df[col].replace(9999, np.nan).mean()
                        if not pd.isna(mean_density):
                            month_data['parameters'][f"Mean Air Density {clean_name} (kg/m³)"] = float(mean_density)
                    for col in speed_cols:
                        clean_name = clean_column_name(col)
                        wind_speeds = month_df[col].replace(9999, np.nan).dropna().values
                        if len(wind_speeds) > 0:
                            epf = calculate_energy_pattern_factor(wind_speeds)
                            month_data['parameters'][f"Energy Pattern Factor {clean_name}"] = float(epf)
                    for pc, pc_data in power_curve_dfs.items():
                        rated_power = pc_data['rated_power']
                        for col in speed_cols:
                            height, _ = get_height_from_column(col)
                            if height >= 140:
                                clean_name = clean_column_name(col)
                                month_df = month_df.copy()
                                wind_speeds = month_df[col].replace(9999, np.nan).values
                                power_output = batch_power_interpolation(
                                    pc_data['wind_speeds'],
                                    pc_data['power_outputs'],
                                    wind_speeds
                                )
                                month_avg_power = np.nanmean(power_output)
                                if month == 2 and year % 4 == 0 and (year % 100 != 0 or year % 400 == 0):
                                    days_in_month = 29
                                elif month == 2:
                                    days_in_month = 28
                                elif month in [4, 6, 9, 11]:
                                    days_in_month = 30
                                else:
                                    days_in_month = 31
                                total_hours = days_in_month * 24
                                gross_energy = month_avg_power * total_hours
                                net_energy = apply_loss_factors(gross_energy)
                                gross_plf = (gross_energy / (rated_power * total_hours)) * 100 if rated_power > 0 and total_hours > 0 else 0
                                net_plf = (net_energy / (rated_power * total_hours)) * 100 if rated_power > 0 and total_hours > 0 else 0
                                month_data['energy'].append({
                                    'power_curve': pc,
                                    'height': clean_name,
                                    'avg_power': float(month_avg_power),
                                    'gross_energy': float(gross_energy),
                                    'net_energy': float(net_energy),
                                    'gross_plf': float(gross_plf),
                                    'net_plf': float(net_plf)
                                })
                    all_results.append(month_data)
            except Exception as e:
                log_error("DATA_SHEET_CHUNK_ERROR", f"Error processing chunk: {str(e)}")
            finally:
                del chunk
                gc.collect()
        if not all_results:
            raise ValueError("No valid data after processing")
        combined_results = []
        results_dict = {}
        for result in all_results:
            key = (result['year'], result['month'])
            if key not in results_dict:
                results_dict[key] = {
                    'year': result['year'],
                    'month': result['month'],
                    'month_name': result['month_name'],
                    'parameters': {},
                    'energy': []
                }
            for param, value in result['parameters'].items():
                if param not in results_dict[key]['parameters']:
                    results_dict[key]['parameters'][param] = []
                results_dict[key]['parameters'][param].append(value)
            results_dict[key]['energy'].extend(result['energy'])
        for key, data in results_dict.items():
            avg_parameters = {}
            for param, values in data['parameters'].items():
                valid_values = [v for v in values if not pd.isna(v)]
                if valid_values:
                    avg_parameters[param] = np.mean(valid_values)
                else:
                    avg_parameters[param] = np.nan
            energy_data = {}
            for e in data['energy']:
                energy_key = f"{e['power_curve']}_{e['height']}"
                if energy_key not in energy_data:
                    energy_data[energy_key] = {
                        'power_curve': e['power_curve'],
                        'height': e['height'],
                        'avg_power': [],
                        'gross_energy': [],
                        'net_energy': [],
                        'gross_plf': [],
                        'net_plf': []
                    }
                energy_data[energy_key]['avg_power'].append(e['avg_power'])
                energy_data[energy_key]['gross_energy'].append(e['gross_energy'])
                energy_data[energy_key]['net_energy'].append(e['net_energy'])
                energy_data[energy_key]['gross_plf'].append(e['gross_plf'])
                energy_data[energy_key]['net_plf'].append(e['net_plf'])
            final_energy = []
            for key, ed in energy_data.items():
                final_energy.append({
                    'power_curve': ed['power_curve'],
                    'height': ed['height'],
                    'avg_power': np.mean(ed['avg_power']),
                    'gross_energy': np.sum(ed['gross_energy']),
                    'net_energy': np.sum(ed['net_energy']),
                    'gross_plf': np.mean(ed['gross_plf']),
                    'net_plf': np.mean(ed['net_plf'])
                })
            combined_results.append({
                'year': data['year'],
                'month': data['month'],
                'month_name': data['month_name'],
                'parameters': avg_parameters,
                'energy': final_energy
            })
        combined_results.sort(key=lambda x: (x['year'], x['month']))
        data_sheet_cache[cache_key] = combined_results
        logger.info(f"Data sheet stats calculated successfully for {state}/{mast} and cached")
        return combined_results
    except Exception as e:
        log_error("DATA_SHEET_ERROR", f"Failed to calculate data sheet stats for {state}/{mast}: {str(e)}")
        return None
def create_excel_data_sheet(data, state, mast, date_range):
    """Optimized Excel data sheet creation with precomputed styles and parallel processing"""
    try:
        start_time = time.time()
        logger.info(f"Creating Excel data sheet for {state}/{mast}")
        if not data:
            logger.error(f"No data provided for Excel creation for {state}/{mast}")
            return None
        cache_key = f"excel_{state}_{mast}_{date_range['type']}_{date_range.get('value', '')}_{date_range.get('start', '')}_{date_range.get('end', '')}"
        cache_file = os.path.join(CACHE_DIR, f"{cache_key}.xlsx")
        if os.path.exists(cache_file):
            logger.info(f"Returning cached Excel file from {cache_file}")
            with open(cache_file, 'rb') as f:
                return base64.b64encode(f.read()).decode('utf-8')
        wb = Workbook()
        ws = wb.active
        ws.title = "Wind Data Analysis"
        styles = {
            'header_fill': PatternFill(start_color="D6E4FF", end_color="D6E4FF", fill_type="solid"),
            'header_font': Font(color="000000", bold=True),
            'center_aligned': Alignment(horizontal="center", vertical="center"),
            'border': Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin')),
            'section_fills': {
                'wind_speed': PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),
                'wind_dir': PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),
                'temp': PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
                'air_density': PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
                'epf': PatternFill(start_color="FFCCFF", end_color="FFCCFF", fill_type="solid"),
                'power': PatternFill(start_color="E6E6FF", end_color="E6E6FF", fill_type="solid"),
                'energy': PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
                'plf': PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
            }
        }
        ws['A1'] = "State:"
        ws['A1'].font = styles['header_font']
        ws['B1'] = state
        ws['C1'] = "Mast:"
        ws['C1'].font = styles['header_font']
        ws['D1'] = mast
        ws['A2'] = "Date Range:"
        ws['A2'].font = styles['header_font']
        if date_range['type'] == 'single_day':
            ws['B2'] = date_range['value']
        elif date_range['type'] == 'monthly':
            ws['B2'] = date_range['value']
        elif date_range['type'] == 'yearly':
            ws['B2'] = date_range['value']
        elif date_range['type'] == 'date_range':
            ws['B2'] = f"{date_range['start']} to {date_range['end']}"
        else:
            ws['B2'] = "Average of all available data"
        yearly_data = defaultdict(list)
        for entry in data:
            if entry and entry.get('month') != 13:  
                yearly_data[entry['year']].append(entry)
        if date_range['type'] == 'average' and yearly_data:
            avg_entries = []
            all_months_data = defaultdict(list)
            for year, months_data in yearly_data.items():
                for month_data in months_data:
                    month = month_data['month']
                    all_months_data[month].append(month_data)
            for month in range(1, 13):
                month_name = datetime(2000, month, 1).strftime('%B')
                month_entries = all_months_data.get(month, [])
                if not month_entries:
                    continue
                aggregated = {
                    'year': 9999,  
                    'month': month,
                    'month_name': month_name,
                    'parameters': defaultdict(list),
                    'energy': defaultdict(lambda: defaultdict(list))
                }
                for entry in month_entries:
                    for param, value in entry['parameters'].items():
                        if not pd.isna(value):
                            aggregated['parameters'][param].append(value)
                    for energy in entry['energy']:
                        key = f"{energy['power_curve']}_{energy['height']}"
                        aggregated['energy'][key]['power_curve'] = energy['power_curve']
                        aggregated['energy'][key]['height'] = energy['height']
                        aggregated['energy'][key]['avg_power'].append(energy['avg_power'])
                        aggregated['energy'][key]['gross_energy'].append(energy['gross_energy'])
                        aggregated['energy'][key]['net_energy'].append(energy['net_energy'])
                        aggregated['energy'][key]['gross_plf'].append(energy['gross_plf'])
                        aggregated['energy'][key]['net_plf'].append(energy['net_plf'])
                final_parameters = {}
                for param, values in aggregated['parameters'].items():
                    if values:
                        final_parameters[param] = np.mean(values)
                final_energy = []
                for key, data in aggregated['energy'].items():
                    final_energy.append({
                        'power_curve': data['power_curve'],
                        'height': data['height'],
                        'avg_power': np.mean(data['avg_power']) if data['avg_power'] else 0,
                        'gross_energy': np.mean(data['gross_energy']) if data['gross_energy'] else 0,
                        'net_energy': np.mean(data['net_energy']) if data['net_energy'] else 0,
                        'gross_plf': np.mean(data['gross_plf']) if data['gross_plf'] else 0,
                        'net_plf': np.mean(data['net_plf']) if data['net_plf'] else 0
                    })
                avg_entries.append({
                    'year': aggregated['year'],
                    'month': aggregated['month'],
                    'month_name': aggregated['month_name'],
                    'parameters': final_parameters,
                    'energy': final_energy
                })
            yearly_data[9999] = avg_entries
        current_row = 4  
        for year in sorted(yearly_data.keys()):
            year_entries = yearly_data[year]
            if year == 9999:
                ws[f'A{current_row}'] = "Average of All Years"
            else:
                ws[f'A{current_row}'] = f"Year: {year}"
            ws[f'A{current_row}'].font = Font(bold=True, size=14)
            current_row += 1
            for i, month in enumerate(MONTHS):
                col = chr(66 + i)  
                ws[f'{col}{current_row}'] = month
                ws[f'{col}{current_row}'].font = styles['header_font']
                ws[f'{col}{current_row}'].alignment = styles['center_aligned']
                ws[f'{col}{current_row}'].fill = styles['header_fill']
                ws[f'{col}{current_row}'].border = styles['border']
            current_row += 1
            ws[f'A{current_row}'] = "Mean Wind Speed (m/s)"
            ws[f'A{current_row}'].fill = styles['section_fills']['wind_speed']
            ws[f'A{current_row}'].font = styles['header_font']
            ws[f'A{current_row}'].border = styles['border']
            current_row += 1
            speed_cols = []
            if year_entries and year_entries[0].get('parameters'):
                for param in year_entries[0]['parameters'].keys():
                    if "Mean Wind Speed" in param:
                        height_dir = param.replace("Mean Wind Speed ", "").replace(" (m/s)", "")
                        speed_cols.append(height_dir)
            speed_cols_sorted = sorted(speed_cols, key=lambda x: int(x.split()[0])) if speed_cols else []
            for height_dir in speed_cols_sorted:
                ws[f'A{current_row}'] = height_dir
                ws[f'A{current_row}'].fill = styles['section_fills']['wind_speed']
                ws[f'A{current_row}'].alignment = styles['center_aligned']
                ws[f'A{current_row}'].border = styles['border']
                values = []
                for i, month in enumerate(MONTHS[:-1]):
                    col = chr(66 + i)
                    param_name = f"Mean Wind Speed {height_dir} (m/s)"
                    if year == 9999:
                        month_data = next((d for d in year_entries if d and d.get('month') == i+1), None)
                    else:
                        month_data = next((d for d in year_entries if d and d.get('month_name') == month), None)
                    if month_data and month_data.get('parameters') and param_name in month_data['parameters']:
                        val = month_data['parameters'][param_name]
                        if not pd.isna(val):
                            ws[f'{col}{current_row}'] = round(val, 2)
                            values.append(val)
                        else:
                            ws[f'{col}{current_row}'] = '-'
                    else:
                        ws[f'{col}{current_row}'] = '-'
                    ws[f'{col}{current_row}'].alignment = styles['center_aligned']
                    ws[f'{col}{current_row}'].border = styles['border']
                if values:
                    avg_col = chr(66 + len(MONTHS) - 1)
                    avg_value = sum(values) / len(values)
                    ws[f'{avg_col}{current_row}'] = round(avg_value, 2)
                    ws[f'{avg_col}{current_row}'].font = styles['header_font']
                    ws[f'{avg_col}{current_row}'].alignment = styles['center_aligned']
                    ws[f'{avg_col}{current_row}'].border = styles['border']
                current_row += 1
            current_row += 1
            ws[f'A{current_row}'] = "Mean Wind Direction (°)"
            ws[f'A{current_row}'].fill = styles['section_fills']['wind_dir']
            ws[f'A{current_row}'].font = styles['header_font']
            ws[f'A{current_row}'].border = styles['border']
            current_row += 1
            dir_cols = []
            if year_entries and year_entries[0].get('parameters'):
                for param in year_entries[0]['parameters'].keys():
                    if "Mean Wind Direction" in param:
                        height_dir = param.replace("Mean Wind Direction ", "").replace(" (°)", "")
                        dir_cols.append(height_dir)
            dir_cols_sorted = sorted(dir_cols, key=lambda x: int(x.split()[0])) if dir_cols else []
            for height_dir in dir_cols_sorted:
                ws[f'A{current_row}'] = height_dir
                ws[f'A{current_row}'].fill = styles['section_fills']['wind_dir']
                ws[f'A{current_row}'].alignment = styles['center_aligned']
                ws[f'A{current_row}'].border = styles['border']
                values = []
                for i, month in enumerate(MONTHS[:-1]):
                    col = chr(66 + i)
                    param_name = f"Mean Wind Direction {height_dir} (°)"
                    if year == 9999:
                        month_data = next((d for d in year_entries if d and d.get('month') == i+1), None)
                    else:
                        month_data = next((d for d in year_entries if d and d.get('month_name') == month), None)
                    if month_data and param_name in month_data['parameters']:
                        val = month_data['parameters'][param_name]
                        if not pd.isna(val):
                            ws[f'{col}{current_row}'] = round(val, 2)
                            values.append(val)
                        else:
                            ws[f'{col}{current_row}'] = '-'
                    else:
                        ws[f'{col}{current_row}'] = '-'
                    ws[f'{col}{current_row}'].alignment = styles['center_aligned']
                    ws[f'{col}{current_row}'].border = styles['border']
                if values:
                    avg_col = chr(66 + len(MONTHS) - 1)
                    sin_sum = sum(np.sin(np.radians(values)))
                    cos_sum = sum(np.cos(np.radians(values)))
                    avg_dir = np.degrees(np.arctan2(sin_sum, cos_sum)) % 360
                    ws[f'{avg_col}{current_row}'] = round(avg_dir, 2)
                    ws[f'{avg_col}{current_row}'].font = styles['header_font']
                    ws[f'{avg_col}{current_row}'].alignment = styles['center_aligned']
                    ws[f'{avg_col}{current_row}'].border = styles['border']
                current_row += 1
            current_row += 1
            ws[f'A{current_row}'] = "Max Temperature (°C)"
            ws[f'A{current_row}'].fill = styles['section_fills']['temp']
            ws[f'A{current_row}'].font = styles['header_font']
            ws[f'A{current_row}'].border = styles['border']
            current_row += 1
            temp_cols = []
            if year_entries and year_entries[0].get('parameters'):
                for param in year_entries[0]['parameters'].keys():
                    if "Max Temperature" in param:
                        height = param.replace("Max Temperature ", "").replace(" (°C)", "")
                        temp_cols.append(height)
            for height in temp_cols:
                ws[f'A{current_row}'] = height
                ws[f'A{current_row}'].fill = styles['section_fills']['temp']
                ws[f'A{current_row}'].alignment = styles['center_aligned']
                ws[f'A{current_row}'].border = styles['border']
                values = []
                for i, month in enumerate(MONTHS[:-1]):
                    col = chr(66 + i)
                    param_name = f"Max Temperature {height} (°C)"
                    if year == 9999:
                        month_data = next((d for d in year_entries if d and d.get('month') == i+1), None)
                    else:
                        month_data = next((d for d in year_entries if d and d.get('month_name') == month), None)
                    if month_data and param_name in month_data['parameters']:
                        val = month_data['parameters'][param_name]
                        if not pd.isna(val):
                            ws[f'{col}{current_row}'] = round(val, 2)
                            values.append(val)
                        else:
                            ws[f'{col}{current_row}'] = '-'
                    else:
                        ws[f'{col}{current_row}'] = '-'
                    ws[f'{col}{current_row}'].alignment = styles['center_aligned']
                    ws[f'{col}{current_row}'].border = styles['border']
                if values:
                    avg_col = chr(66 + len(MONTHS) - 1)
                    avg_value = sum(values) / len(values)
                    ws[f'{avg_col}{current_row}'] = round(avg_value, 2)
                    ws[f'{avg_col}{current_row}'].font = styles['header_font']
                    ws[f'{avg_col}{current_row}'].alignment = styles['center_aligned']
                    ws[f'{avg_col}{current_row}'].border = styles['border']
                current_row += 1
            current_row += 1
            ws[f'A{current_row}'] = "Mean Air Density (kg/m³)"
            ws[f'A{current_row}'].fill = styles['section_fills']['air_density']
            ws[f'A{current_row}'].font = styles['header_font']
            ws[f'A{current_row}'].border = styles['border']
            current_row += 1
            air_density_cols = []
            if year_entries and year_entries[0].get('parameters'):
                for param in year_entries[0]['parameters'].keys():
                    if "Mean Air Density" in param:
                        height = param.replace("Mean Air Density ", "").replace(" (kg/m³)", "")
                        air_density_cols.append(height)
            for height in air_density_cols:
                ws[f'A{current_row}'] = height
                ws[f'A{current_row}'].fill = styles['section_fills']['air_density']
                ws[f'A{current_row}'].alignment = styles['center_aligned']
                ws[f'A{current_row}'].border = styles['border']
                values = []
                for i, month in enumerate(MONTHS[:-1]):
                    col = chr(66 + i)
                    param_name = f"Mean Air Density {height} (kg/m³)"
                    if year == 9999:
                        month_data = next((d for d in year_entries if d and d.get('month') == i+1), None)
                    else:
                        month_data = next((d for d in year_entries if d and d.get('month_name') == month), None)
                    if month_data and param_name in month_data['parameters']:
                        val = month_data['parameters'][param_name]
                        if not pd.isna(val):
                            ws[f'{col}{current_row}'] = round(val, 3)
                            values.append(val)
                        else:
                            ws[f'{col}{current_row}'] = '-'
                    else:
                        ws[f'{col}{current_row}'] = '-'
                    ws[f'{col}{current_row}'].alignment = styles['center_aligned']
                    ws[f'{col}{current_row}'].border = styles['border']
                if values:
                    avg_col = chr(66 + len(MONTHS) - 1)
                    avg_value = sum(values) / len(values)
                    ws[f'{avg_col}{current_row}'] = round(avg_value, 3)
                    ws[f'{avg_col}{current_row}'].font = styles['header_font']
                    ws[f'{avg_col}{current_row}'].alignment = styles['center_aligned']
                    ws[f'{avg_col}{current_row}'].border = styles['border']
                current_row += 1
            current_row += 1
            ws[f'A{current_row}'] = "Energy Pattern Factor"
            ws[f'A{current_row}'].fill = styles['section_fills']['epf']
            ws[f'A{current_row}'].font = styles['header_font']
            ws[f'A{current_row}'].border = styles['border']
            current_row += 1
            epf_cols = []
            if year_entries and year_entries[0].get('parameters'):
                for param in year_entries[0]['parameters'].keys():
                    if "Energy Pattern Factor" in param:
                        height_dir = param.replace("Energy Pattern Factor ", "")
                        epf_cols.append(height_dir)
            epf_cols_sorted = sorted(epf_cols, key=lambda x: int(x.split()[0])) if epf_cols else []
            for height_dir in epf_cols_sorted:
                ws[f'A{current_row}'] = height_dir
                ws[f'A{current_row}'].fill = styles['section_fills']['epf']
                ws[f'A{current_row}'].alignment = styles['center_aligned']
                ws[f'A{current_row}'].border = styles['border']
                values = []
                for i, month in enumerate(MONTHS[:-1]):
                    col = chr(66 + i)
                    param_name = f"Energy Pattern Factor {height_dir}"
                    if year == 9999:
                        month_data = next((d for d in year_entries if d and d.get('month') == i+1), None)
                    else:
                        month_data = next((d for d in year_entries if d and d.get('month_name') == month), None)
                    if month_data and param_name in month_data['parameters']:
                        val = month_data['parameters'][param_name]
                        if not pd.isna(val):
                            ws[f'{col}{current_row}'] = round(val, 3)
                            values.append(val)
                        else:
                            ws[f'{col}{current_row}'] = '-'
                    else:
                        ws[f'{col}{current_row}'] = '-'
                    ws[f'{col}{current_row}'].alignment = styles['center_aligned']
                    ws[f'{col}{current_row}'].border = styles['border']
                if values:
                    avg_col = chr(66 + len(MONTHS) - 1)
                    avg_value = sum(values) / len(values)
                    ws[f'{avg_col}{current_row}'] = round(avg_value, 3)
                    ws[f'{avg_col}{current_row}'].font = styles['header_font']
                    ws[f'{avg_col}{current_row}'].alignment = styles['center_aligned']
                    ws[f'{avg_col}{current_row}'].border = styles['border']
                current_row += 1
            current_row += 1
            ws[f'A{current_row}'] = "Mean Power Output (kW)"
            ws[f'A{current_row}'].fill = styles['section_fills']['power']
            ws[f'A{current_row}'].font = styles['header_font']
            ws[f'A{current_row}'].border = styles['border']
            current_row += 1
            power_curves = set()
            heights = set()
            for month_data in year_entries:
                for energy in month_data['energy']:
                    power_curves.add(energy['power_curve'])
                    height = energy['height'].split()[0]
                    if int(height) >= 140:
                        heights.add(energy['height'])
            for pc in sorted(power_curves):
                for height in sorted(heights, key=lambda x: int(x.split()[0])):
                    ws[f'A{current_row}'] = f"{pc} {height}"
                    ws[f'A{current_row}'].fill = styles['section_fills']['power']
                    ws[f'A{current_row}'].alignment = styles['center_aligned']
                    ws[f'A{current_row}'].border = styles['border']
                    values = []
                    for i, month in enumerate(MONTHS[:-1]):
                        col = chr(66 + i)
                        if year == 9999:
                            month_data = next((d for d in year_entries if d and d.get('month') == i+1), None)
                        else:
                            month_data = next((d for d in year_entries if d and d.get('month_name') == month), None)
                        if month_data:
                            energy_data = next((e for e in month_data['energy'] 
                                            if e['power_curve'] == pc and e['height'] == height), None)
                            if energy_data:
                                ws[f'{col}{current_row}'] = round(energy_data['avg_power'], 2)
                                values.append(energy_data['avg_power'])
                            else:
                                ws[f'{col}{current_row}'] = '-'
                        else:
                            ws[f'{col}{current_row}'] = '-'
                        ws[f'{col}{current_row}'].alignment = styles['center_aligned']
                        ws[f'{col}{current_row}'].border = styles['border']
                    if values:
                        avg_col = chr(66 + len(MONTHS) - 1)
                        avg_value = sum(values) / len(values)
                        ws[f'{avg_col}{current_row}'] = round(avg_value, 2)
                        ws[f'{avg_col}{current_row}'].font = styles['header_font']
                        ws[f'{avg_col}{current_row}'].alignment = styles['center_aligned']
                        ws[f'{avg_col}{current_row}'].border = styles['border']
                    current_row += 1
            current_row += 1
            ws[f'A{current_row}'] = "Net Energy (kWh)"
            ws[f'A{current_row}'].fill = styles['section_fills']['energy']
            ws[f'A{current_row}'].font = styles['header_font']
            ws[f'A{current_row}'].border = styles['border']
            current_row += 1
            for pc in sorted(power_curves):
                for height in sorted(heights, key=lambda x: int(x.split()[0])):
                    ws[f'A{current_row}'] = f"{pc} {height}"
                    ws[f'A{current_row}'].fill = styles['section_fills']['energy']
                    ws[f'A{current_row}'].alignment = styles['center_aligned']
                    ws[f'A{current_row}'].border = styles['border']
                    values = []
                    for i, month in enumerate(MONTHS[:-1]):
                        col = chr(66 + i)
                        if year == 9999:
                            month_data = next((d for d in year_entries if d and d.get('month') == i+1), None)
                        else:
                            month_data = next((d for d in year_entries if d and d.get('month_name') == month), None)
                        if month_data:
                            energy_data = next((e for e in month_data['energy'] 
                                              if e['power_curve'] == pc and e['height'] == height), None)
                            if energy_data:
                                ws[f'{col}{current_row}'] = round(energy_data['net_energy'], 2)
                                values.append(energy_data['net_energy'])
                            else:
                                ws[f'{col}{current_row}'] = '-'
                        else:
                            ws[f'{col}{current_row}'] = '-'
                        ws[f'{col}{current_row}'].alignment = styles['center_aligned']
                        ws[f'{col}{current_row}'].border = styles['border']
                    if values:
                        avg_col = chr(66 + len(MONTHS) - 1)
                        avg_value = sum(values) / len(values)
                        ws[f'{avg_col}{current_row}'] = round(avg_value, 2)
                        ws[f'{avg_col}{current_row}'].font = styles['header_font']
                        ws[f'{avg_col}{current_row}'].alignment = styles['center_aligned']
                        ws[f'{avg_col}{current_row}'].border = styles['border']
                    current_row += 1
            current_row += 1
            ws[f'A{current_row}'] = "Net PLF (%)"
            ws[f'A{current_row}'].fill = styles['section_fills']['plf']
            ws[f'A{current_row}'].font = styles['header_font']
            ws[f'A{current_row}'].border = styles['border']
            current_row += 1
            for pc in sorted(power_curves):
                for height in sorted(heights, key=lambda x: int(x.split()[0])):
                    ws[f'A{current_row}'] = f"{pc} {height}"
                    ws[f'A{current_row}'].fill = styles['section_fills']['plf']
                    ws[f'A{current_row}'].alignment = styles['center_aligned']
                    ws[f'A{current_row}'].border = styles['border']
                    values = []
                    for i, month in enumerate(MONTHS[:-1]):
                        col = chr(66 + i)
                        if year == 9999:
                            month_data = next((d for d in year_entries if d and d.get('month') == i+1), None)
                        else:
                            month_data = next((d for d in year_entries if d and d.get('month_name') == month), None)
                        if month_data:
                            energy_data = next((e for e in month_data['energy'] 
                                            if e['power_curve'] == pc and e['height'] == height), None)
                            if energy_data:
                                ws[f'{col}{current_row}'] = round(energy_data['net_plf'], 2)
                                values.append(energy_data['net_plf'])
                            else:
                                ws[f'{col}{current_row}'] = '-'
                        else:
                            ws[f'{col}{current_row}'] = '-'
                        ws[f'{col}{current_row}'].alignment = styles['center_aligned']
                        ws[f'{col}{current_row}'].border = styles['border']
                    if values:
                        avg_col = chr(66 + len(MONTHS) - 1)
                        avg_value = sum(values) / len(values)
                        ws[f'{avg_col}{current_row}'] = round(avg_value, 2)
                        ws[f'{avg_col}{current_row}'].font = styles['header_font']
                        ws[f'{avg_col}{current_row}'].alignment = styles['center_aligned']
                        ws[f'{avg_col}{current_row}'].border = styles['border']
                    current_row += 1
        ws.column_dimensions['A'].width = 30
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws.column_dimensions[col].width = 15
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        with open(cache_file, 'wb') as f:
            f.write(output.getvalue())
        logger.info(f"Excel data sheet created in {time.time() - start_time:.2f} seconds and cached")
        return base64.b64encode(output.read()).decode('utf-8')
    except Exception as e:
        logger.error(f"Error creating Excel data sheet: {str(e)}")
        traceback.print_exc()
        return None
def excel_preparation_worker():
    """Background worker to prepare Excel sheets in parallel"""
    while True:
        try:
            task = excel_preparation_queue.get()
            if task is None:  
                break
            state, mast, date_range, data = task
            cache_key = f"excel_{state}_{mast}_{date_range['type']}_{date_range.get('value', '')}_{date_range.get('start', '')}_{date_range.get('end', '')}"
            with excel_preparation_lock:
                if cache_key not in excel_preparation_cache:
                    logger.info(f"Background worker preparing Excel for {state}/{mast}")
                    if data is None:
                        power_curves = ['suzlon']  
                        data = calculate_data_sheet_stats(state, mast, date_range, power_curves)
                    if data:
                        excel_data = create_excel_data_sheet(data, state, mast, date_range)
                        if excel_data:
                            excel_preparation_cache[cache_key] = excel_data
        except Exception as e:
            logger.error(f"Error in Excel preparation worker: {str(e)}")
        finally:
            excel_preparation_queue.task_done()
NUM_EXCEL_WORKERS = 4
excel_threads = []
for i in range(NUM_EXCEL_WORKERS):
    t = threading.Thread(target=excel_preparation_worker, daemon=True)
    t.start()
    excel_threads.append(t)
@app.route('/energy')
def modi():
    states = ['Andhra Pradesh', 'Madhya Pradesh', 'Rajasthan', 'Telangana']
    power_curves = get_power_curves()
    return render_template('energy.html', states=states, power_curves=power_curves)
@app.route('/get_masts', methods=['POST'])
def get_masts():
    state = request.json['state']
    return jsonify({'masts': get_state_masts(state)})
@app.route('/get_columns', methods=['POST'])
def get_columns():
    state = request.json['state']
    mast = request.json['mast']
    return jsonify({'columns': get_mast_columns(state, mast)})
@app.route('/validate_file', methods=['POST'])
def validate_file_endpoint():
    try:
        data = request.json
        state = data['state']
        mast = data['mast']
        return jsonify(validate_file(state, mast))
    except Exception as e:
        return jsonify({'valid': False, 'error': str(e)})
@app.route('/extrapolate', methods=['POST'])
def handle_extrapolate():
    try:
        data = request.json
        state = data['state']
        mast = data['mast']
        speed_col = data['speed_col']
        target_height = float(data['target_height'])
        selected_heights = data.get('selected_heights', [])
        extrapolate_from = data.get('extrapolate_from', speed_col)
        file_path = os.path.join(DATA_DIR, state, f"{mast}.csv")
        chunksize = 100000
        extrapolated_dict = {}
        valid_count = 0
        for chunk in pd.read_csv(file_path, chunksize=chunksize):
            for col in chunk.columns:
                if 'Spd' in col:
                    chunk[col] = pd.to_numeric(chunk[col], errors='coerce')
                    chunk[col] = chunk[col].replace(9999, np.nan)
            alpha = None
            if len(selected_heights) == 2:
                try:
                    h1 = float(selected_heights[0].split(' ')[1].replace('m', ''))
                    h2 = float(selected_heights[1].split(' ')[1].replace('m', ''))
                    v1 = chunk[selected_heights[0]].mean()
                    v2 = chunk[selected_heights[1]].mean()
                    alpha = calculate_wind_shear_numba(v1, v2, h1, h2)
                except Exception as e:
                    logger.error(f"Error calculating alpha: {str(e)}")
            new_col = extrapolate_wind_speeds(chunk, extrapolate_from, target_height, alpha)
            if not new_col:
                continue
            timestamp_col = next((col for col in chunk.columns if 'Timestamp' in col), None)
            if not timestamp_col:
                continue
            timestamps = chunk[timestamp_col].astype(str)
            speeds = chunk[new_col].values
            for ts, speed in zip(timestamps, speeds):
                if pd.notna(speed):
                    try:
                        if ' ' in ts:
                            standardized_ts = ts
                        else:
                            dt = pd.to_datetime(ts, errors='coerce')
                            if pd.notna(dt):
                                standardized_ts = dt.strftime('%d%m%Y %H%M')
                            else:
                                continue
                        if standardized_ts not in extrapolated_dict:
                            extrapolated_dict[standardized_ts] = float(speed)
                            valid_count += 1
                    except Exception as e:
                        logger.error(f"Error processing timestamp {ts}: {str(e)}")
                        continue
        key = f"{state}_{mast}_{new_col}"
        extrapolated_data[key] = extrapolated_dict
        if not save_extrapolated_data():
            return jsonify({'success': False, 'error': 'Failed to save extrapolated data'})
        return jsonify({
            'success': True,
            'alpha': alpha if alpha is not None else 1/7,
            'new_column': new_col,
            'data_points': valid_count
        })
    except Exception as e:
        logger.error(f"Error in extrapolation endpoint: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})
@app.route('/process', methods=['POST'])
def process():
    """Endpoint with enhanced error handling, logging, and parallel processing"""
    try:
        start_time = time.time()
        data = request.json
        logger.info("Received process request")
        if not data:
            logger.error("No data received in request")
            return jsonify({'success': False, 'error': 'No data received'}), 400
        date_range = data.get('date_range', {})
        selected_sites = data.get('selected_sites', [])
        power_curves = data.get('power_curves', [])
        if not all([date_range, selected_sites, power_curves]):
            logger.error("Missing required parameters")
            return jsonify({'success': False, 'error': 'Missing required parameters'}), 400
        results = []
        data_sheet_results = {}
        errors = []
        max_workers = min(8, cpu_count() * 2)
        logger.info(f"Using {max_workers} workers for parallel processing")
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = []
            for site in selected_sites:
                for pc in power_curves:
                    futures.append(executor.submit(
                        process_data,
                        site['state'],
                        site['mast'],
                        site['speed_col'],
                        pc,
                        date_range
                    ))
                    futures.append(executor.submit(
                        calculate_data_sheet_stats,
                        site['state'],
                        site['mast'],
                        date_range,
                        power_curves
                    ))
            for site in selected_sites:
                excel_preparation_queue.put((
                    site['state'],
                    site['mast'],
                    date_range,
                    None  
                ))
            for future in as_completed(futures, timeout=300):  
                try:
                    result = future.result()
                    if result:
                        if 'monthly_stats' in result:  
                            results.append(result)
                        else:  
                            site_key = f"{result[0]['state']}_{result[0]['mast']}" if result else None
                            if site_key:
                                data_sheet_results[site_key] = result
                                excel_preparation_queue.put((
                                    result[0]['state'],
                                    result[0]['mast'],
                                    date_range,
                                    result
                                ))
                except Exception as e:
                    errors.append({'error': str(e)})
                    logger.error(f"Error processing site: {str(e)}")
        logger.info(f"Processing completed in {time.time() - start_time:.2f} seconds with {len(results)} results and {len(errors)} errors")
        def convert_numpy_types(obj):
            if isinstance(obj, (np.integer, np.int32, np.int64)):
                return int(obj)
            elif isinstance(obj, (np.floating, np.float32, np.float64)):
                return float(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, dict):
                return {k: convert_numpy_types(v) for k, v in obj.items()}
            elif isinstance(obj, (list, tuple)):
                return [convert_numpy_types(item) for item in obj]
            return obj
        response_data = {
            'success': bool(results),
            'results': convert_numpy_types(results),
            'data_sheet_results': convert_numpy_types(data_sheet_results),
            'errors': convert_numpy_types(errors) if errors else None,
            'error_summary': convert_numpy_types({
                'last_errors': error_tracker['last_errors'][-5:],
                'error_counts': error_tracker['error_counts']
            })
        }
        return jsonify(response_data)
    except Exception as e:
        log_error("PROCESS_ENDPOINT_ERROR", f"Failed to handle process request: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'error_details': error_tracker['last_errors'][-5:]
        }), 500
@app.route('/get_data_sheet', methods=['POST'])
def get_data_sheet():
    """Endpoint for data sheet with proper error handling"""
    try:
        data = request.json
        logger.info("Received data sheet request")
        if not data:
            return jsonify({'success': False, 'error': 'No data received'}), 400
        state = data.get('state')
        mast = data.get('mast')
        date_range = data.get('date_range', {})
        power_curves = data.get('power_curves', [])
        if not all([state, mast, date_range, power_curves]):
            return jsonify({'success': False, 'error': 'Missing required parameters'}), 400
        results = calculate_data_sheet_stats(state, mast, date_range, power_curves)
        if not results:
            return jsonify({'success': False, 'error': 'No valid data found'})
        return jsonify({
            'success': True, 
            'data': results,
            'filename': f"{state}_{mast}_data_sheet.json"
        })
    except Exception as e:
        log_error("DATA_SHEET_ENDPOINT_ERROR", f"Failed to handle data sheet request: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'error_details': error_tracker['last_errors'][-3:]
        }), 500
@app.route('/download_data_sheet', methods=['POST'])
def download_data_sheet():
    try:
        data = request.json
        if 'date_range' in data:
            date_range = data['date_range']
        else:
            date_range = {
                'type': data.get('date_range_type', 'single_day')
            }
            if date_range['type'] == 'single_day':
                date_range['value'] = data.get('single_date', '')
            elif date_range['type'] == 'monthly':
                date_range['value'] = data.get('monthly_date', '')
            elif date_range['type'] == 'yearly':
                date_range['value'] = data.get('yearly_date', '')
            elif date_range['type'] == 'date_range':
                date_range['start'] = data.get('start_date', '')
                date_range['end'] = data.get('end_date', '')
        state = data['state']
        mast = data['mast']
        cache_key = f"excel_{state}_{mast}_{date_range['type']}_{date_range.get('value', '')}_{date_range.get('start', '')}_{date_range.get('end', '')}"
        with excel_preparation_lock:
            if cache_key in excel_preparation_cache:
                logger.info(f"Returning pre-prepared Excel for {cache_key}")
                excel_data = excel_preparation_cache[cache_key]
            else:
                logger.info(f"No pre-prepared Excel found for {cache_key}, generating now")
                if 'data' in data:
                    results = data['data']
                else:
                    power_curves = data.get('power_curves', [])
                    results = calculate_data_sheet_stats(state, mast, date_range, power_curves)
                if not results:
                    return jsonify({'success': False, 'error': 'No valid data found for the selected site'})
                excel_data = create_excel_data_sheet(results, state, mast, date_range)
                if not excel_data:
                    return jsonify({'success': False, 'error': 'Error creating Excel file'})
        return jsonify({
            'success': True, 
            'file_data': excel_data,
            'filename': f"{state}_{mast}_wind_data_analysis.xlsx"
        })
    except Exception as e:
        logger.error(f"Error in download_data_sheet endpoint: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})
@app.route('/update_loss_factors', methods=['POST'])
def update_loss_factors():
    global loss_factors
    try:
        data = request.json
        loss_factors.update(data)
        update_total_loss_factor()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error updating loss factors: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})
@app.route('/clear_extrapolated_data', methods=['POST'])
def clear_extrapolated_data_endpoint():
    global extrapolated_data
    try:
        extrapolated_data = {}
        if os.path.exists(EXTRAPOLATED_DATA_PATH):
            os.remove(EXTRAPOLATED_DATA_PATH)
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error clearing extrapolated data: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})
def cleanup():
    for _ in range(NUM_EXCEL_WORKERS):
        excel_preparation_queue.put(None)  
    for t in excel_threads:
        t.join()
if __name__ == '__main__':
    freeze_support()
    os.makedirs(KML_DIR, exist_ok=True)
    logger.info(f"KMZ/KML directory: {KML_DIR}")
    WSGIRequestHandler.protocol_version = "HTTP/1.1"
    app.run(host='0.0.0.0', port=5000, debug=True ,threaded=True)